import pandas as pd
import re
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

def process_assistant_text(text):
    # Keywords to identify where the JSON sequence starts
    keywords = ["Here�s the new sequence in JSON format:", "JSON format", "Here�s the new sequence I can generate for you"]
    for keyword in keywords:
        if keyword in text:
            # Split the content based on the keyword
            parts = text.split(keyword, 1)
            if len(parts) > 1:
                # Add the specific phrase before the JSON
                modified_text = parts[0].strip() + "\n[this part is not spoken by furhat]\n" + keyword + parts[1].strip()
                return modified_text
    return text

def parse_text(text):
    # Split the text into individual lines
    lines = text.strip().split('\n')

    # Variables to store parsed data
    data = []
    participant_id = None

    # Regular expressions to identify the parts of the text
    id_pattern = re.compile(r'Participant (\d+)')
    user_pattern = re.compile(r'^User:')
    assistant_pattern = re.compile(r'^Assistant:')
    system_keywords = [
        "Is it ambiguous?", "Does it map to the action in var_KnowNo?", 
        "Furhat is being used with Baseline", "var_KnowNo", 
        "Tree is initialized", "Run number", "This run took", "Total number of LLM calls"
    ]

    # Iterate through the lines to categorize and extract information
    current_speaker = None
    current_text = []

    for line in lines:
        # Check for a new participant ID
        match = id_pattern.search(line)
        if match:
            participant_id = match.group(1)
            continue  # Skip to the next line after extracting the participant ID

        if user_pattern.match(line):
            if current_speaker:
                if current_speaker == 'Assistant':
                    # Process and append the entire assistant output before adding it to the data
                    ass_output = "\n".join(current_text).strip()
                    ass_output = process_assistant_text(ass_output)
                    data.append((participant_id, current_speaker, ass_output))
                else:
                    data.append((participant_id, current_speaker, "\n".join(current_text).strip()))
            current_speaker = 'User'
            # Check if the split result has the expected format
            parts = line.split('User: ', 1)
            if len(parts) > 1:
                current_text = [parts[1].strip()]
            else:
                current_text = []
        elif assistant_pattern.match(line):
            if current_speaker:
                if current_speaker == 'Assistant':
                    # Process and append the entire assistant output before adding it to the data
                    ass_output = "\n".join(current_text).strip()
                    ass_output = process_assistant_text(ass_output)
                    data.append((participant_id, current_speaker, ass_output))
                else:
                    data.append((participant_id, current_speaker, "\n".join(current_text).strip()))
            current_speaker = 'Assistant'
            parts = line.split('Assistant: ', 1)
            if len(parts) > 1:
                current_text = [parts[1].strip()]
            else:
                current_text = []
        elif any(keyword in line for keyword in system_keywords) or line.startswith("Time:") or line.startswith("Version:") or line.startswith("Run:"):
            if current_speaker:
                if current_speaker == 'Assistant':
                    ass_output = "\n".join(current_text).strip()
                    ass_output = process_assistant_text(ass_output)
                    data.append((participant_id, current_speaker, ass_output))
                else:
                    data.append((participant_id, current_speaker, "\n".join(current_text).strip()))
            current_speaker = 'System'
            current_text = [line.strip()]
        else:
            current_text.append(line.strip())

    if current_speaker:
        if current_speaker == 'Assistant':
            ass_output = "\n".join(current_text).strip()
            ass_output = process_assistant_text(ass_output)
            data.append((participant_id, current_speaker, ass_output))
        else:
            data.append((participant_id, current_speaker, "\n".join(current_text).strip()))

    return data

def write_to_excel(data, filename='Transkript of Conversation/transcript_parsed.xlsx'):
    # Create a DataFrame
    df = pd.DataFrame(data, columns=['ID', 'Who', 'What'])

    try:
        # Load the existing workbook if it exists
        workbook = load_workbook(filename)
        with pd.ExcelWriter(filename, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            # Add the DataFrame to a new sheet called 'Transcript'
            df.to_excel(writer, index=False, sheet_name='Example Transcript')

            # Access the newly created sheet for formatting
            worksheet = writer.sheets['Example Transcript']

            # Apply formatting and merging as needed
            fill_user = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Light green
            fill_assistant = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")  # Light blue
            fill_system = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Light orange

            # Merge cells for the ID column per participant
            start_row = 2
            for row in range(3, len(df) + 2):
                previous_id = worksheet.cell(row=start_row, column=1).value
                current_id = worksheet.cell(row=row, column=1).value

                if current_id != previous_id:
                    worksheet.merge_cells(start_row=start_row, start_column=1, end_row=row - 1, end_column=1)
                    merged_cell = worksheet.cell(row=start_row, column=1)
                    merged_cell.alignment = Alignment(horizontal='center', vertical='top')
                    start_row = row

            if len(df) + 1 > start_row:
                worksheet.merge_cells(start_row=start_row, start_column=1, end_row=len(df) + 1, end_column=1)
                merged_cell = worksheet.cell(row=start_row, column=1)
                merged_cell.alignment = Alignment(horizontal='center', vertical='top')

            # Merge repeated "Who" column values
            start_row = 2
            for row in range(3, len(df) + 2):
                previous_value = worksheet.cell(row=start_row, column=2).value
                current_value = worksheet.cell(row=row, column=2).value

                if current_value != previous_value:
                    if row - 1 > start_row:
                        worksheet.merge_cells(start_row=start_row, start_column=2, end_row=row - 1, end_column=2)
                        merged_cell = worksheet.cell(row=start_row, column=2)
                        merged_cell.value = f"{previous_value}:"
                        merged_cell.alignment = Alignment(vertical='top')
                    start_row = row

            if len(df) + 1 > start_row:
                worksheet.merge_cells(start_row=start_row, start_column=2, end_row=len(df) + 1, end_column=2)
                merged_cell = worksheet.cell(row=start_row, column=2)
                merged_cell.alignment = Alignment(vertical='top')

            for row in range(2, len(df) + 2):
                who_value = worksheet.cell(row=row, column=2).value
                who_cell = worksheet.cell(row=row, column=2)
                what_cell = worksheet.cell(row=row, column=3)

                cell_coord = f"B{row}"
                merged = any(cell_coord in merged_range for merged_range in worksheet.merged_cells.ranges)

                if not merged:
                    who_cell.value = f"{who_value}:"

                if who_value == 'User' or who_value == 'User:':
                    who_cell.fill = fill_user
                elif who_value == 'Assistant' or who_value == 'Assistant:':
                    who_cell.fill = fill_assistant
                elif who_value == 'System' or who_value == 'System:':
                    who_cell.fill = fill_system

                what_cell.alignment = Alignment(wrap_text=True, vertical='top')
                who_cell.alignment = Alignment(vertical='top')

            for col in range(1, 4):
                max_length = 0
                col_letter = get_column_letter(col)
                for cell in worksheet[col_letter]:
                    try:
                        max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                worksheet.column_dimensions[col_letter].width = max_length + 5

    except FileNotFoundError:
        # If the file does not exist, create a new one
        with pd.ExcelWriter(filename, engine='openpyxl', mode='w') as writer:
            df.to_excel(writer, index=False, sheet_name='Example Transcript')

# Example usage
text = """
Participant 43
RunID C
Version: 1.3
Tue Sep 24 10:05:15 2024
Furhat is being used with Baseline
Assistant: Hello! I am Gregory, your home assistant. How can I help you today?
Assistant: I am sorry. I did not get that. Can you please repeat?
User: can you please help me prepare uh bean and cheese kadia 
Time: Tue Sep 24 10:05:41 2024
Assistant: It looks like you might be requesting a "bean and cheese quesadilla," but I'm not familiar with "bean and cheese kadia." 

If you meant "bean and cheese quesadilla," I can assist you with that! I will prepare a quesadilla using heated beans and cheese with a tortilla. 

First, I will heat the beans, prepare the tortilla, and then assemble and cook the quesadilla.

Here�s the new sequence in JSON format: 
```json
{
  "task": "bean and cheese quesadilla",
  "ID": "8",
  "sequence": [
    {"step": "1", "action": "heat_beans(1)"},
    {"step": "2", "action": "prepare_tortilla(1)"},
    {"step": "3", "action": "place_on_plate(prepared_tortilla, 1)"},
    {"step": "4", "action": "place_on_plate(heated_beans, 1)"},
    {"step": "5", "action": "place_on_plate(cheese, 1)"},
    {"step": "6", "action": "cook_quesadilla()"},
    {"step": "7", "action": "place_on_plate(cooked_quesadilla, 1)"},
    {"step": "8", "action": "serve()"}
  ]
}
```
Great choice! I will start the task for: bean and cheese quesadilla.
Time: Tue Sep 24 10:05:47 2024
Run number: 1 ended at: Tue Sep 24 10:05:47 2024
This run took 1 turns
Total number of LLM calls: 1

Version: 1.3
Tue Sep 24 10:06:10 2024
Furhat is being used with Baseline
Assistant: That's it for this task. Let's go onto the next one... Hello. How can I help you today?
User:  uh can you suggest something sweet Savory or healthy from the list 
Time: Tue Sep 24 10:06:30 2024
Assistant: It seems your request is a bit vague. Could you please clarify what you mean by "something sweet, savory or healthy"? Here are a few options from the list that might fit those categories:

1. **Sweet**: Pancakes with maple syrup and berries.
2. **Savory**: Bacon and egg sandwich.
3. **Healthy**: Avocado toast with sausage on the side.

Let me know if any of these options interest you, or if you have something else in mind!
Time: Tue Sep 24 10:06:32 2024
User: uh can you prepare uh avocado toast for me 
Time: Tue Sep 24 10:07:03 2024
Assistant: Great choice! I will start the task for: avocado toast with sausage on the side.
Time: Tue Sep 24 10:07:04 2024
Run number: 2 ended at: Tue Sep 24 10:07:04 2024
This run took 2 turns
Total number of LLM calls: 2

Version: 1.3
Tue Sep 24 10:07:10 2024
Furhat is being used with Baseline
Assistant: That's it for this task. Let's go onto the next one... Hello. How can I help you today?
User:   can you prepare peanut butter and jelly sandwich but with grilled Tomatoes added 
Time: Tue Sep 24 10:07:38 2024
Assistant: First, I will prepare a peanut butter and jelly sandwich and then add grilled tomatoes as an extra ingredient. I will toast the bread, spread peanut butter and jelly, and then place the grilled tomatoes on the sandwich. 

Here�s the new sequence in JSON format: 
```json
{
  "task": "peanut butter and jelly sandwich with grilled tomatoes",
  "ID": "3a",
  "sequence": [
    {"step": "1", "action": "toast_bread(2)"},
    {"step": "2", "action": "spread_on_bread(peanut_butter)"},
    {"step": "3", "action": "spread_on_bread(jelly)"},
    {"step": "4", "action": "grill_tomatoes(1)"},
    {"step": "5", "action": "combine_breads(jelly_toast, peanut_butter_toast)"},
    {"step": "6", "action": "place_on_sandwich(grilled_tomatoes, 1)"},
    {"step": "7", "action": "serve()"}
  ]
}
```
Time: Tue Sep 24 10:07:41 2024
Run number: 3 ended at: Tue Sep 24 10:07:41 2024
This run took 3 turns
Total number of LLM calls: 1

Version: 1.3
Tue Sep 24 10:07:56 2024
Furhat is being used with Baseline
Assistant: That's it for this task. Let's go onto the next one... Hello. How can I help you today?
User:  can you help me uh clean the floor 
Time: Tue Sep 24 10:08:14 2024
Assistant: It seems that your request is a bit vague. However, I can assist with cleaning the living room floor. Here�s how I can help:

1. Clean living room floor 

Please confirm if you would like me to proceed with cleaning the living room floor!
Time: Tue Sep 24 10:08:15 2024
User: yes
Time: Tue Sep 24 10:08:43 2024
Assistant: Great choice! I will start the task for: clean living room floor.
Time: Tue Sep 24 10:08:44 2024
Run number: 4 ended at: Tue Sep 24 10:08:44 2024
This run took 4 turns
Total number of LLM calls: 2

Version: 1.3
Tue Sep 24 10:08:48 2024
Furhat is being used with Baseline
Assistant: That's it for this task. Let's go onto the next one... Hello. How can I help you today?
User: hey uh can I get a a Mexican dish with tortilla and cheese 
Time: Tue Sep 24 10:09:07 2024
Assistant: It seems like you're looking for a Mexican dish that includes tortilla and cheese. I can prepare a "bean and cheese quesadilla" for you, which involves heating beans, preparing a tortilla, and cooking it with cheese.

First, I will heat the beans, prepare the tortilla, add the beans and cheese, then cook the quesadilla until it's crispy. 

Great choice! I will start the task for: bean and cheese quesadilla.
Time: Tue Sep 24 10:09:11 2024
Run number: 5 ended at: Tue Sep 24 10:09:11 2024
This run took 5 turns
Total number of LLM calls: 1

Version: 1.3
Tue Sep 24 10:14:58 2024
Tree is initialized and furhat is used

Assistant: Hello! I am Gregory, your home assistant. How can I help you today?
User: uh can you help me prepare a pancakes with maple syrup and berries 
Is it ambiguous? False
Time: Tue Sep 24 10:14:58 2024
Is it known? True
Time: Tue Sep 24 10:14:59 2024
var_KnowNo: ['pancakes with maple syrup and berries']
Time: Tue Sep 24 10:15:01 2024
Does it map to the action in var_KnowNo? True
Time: Tue Sep 24 10:15:02 2024
Assistant: Great choice! I will get the task for pancakes with maple syrup and berries started.
Time: Tue Sep 24 10:15:02 2024
Run number: 1 ended at: Tue Sep 24 10:15:02 2024
This run took 3 turns
Total number of LLM calls: 5

Run: 2
Tue Sep 24 10:15:21 2024
Assistant: That's it for this task. Let's go onto the next one... Hello. How can I help you today?
User: uh can you suggest something sweet from the list 
Is it ambiguous? True
Time: Tue Sep 24 10:15:21 2024
Assistant: There are multiple options in my list that might fit your order. For something sweet, I recommend 'pancakes with maple syrup and berries'. It's a delightful dish that's perfect for satisfying a sweet craving. Would you like to have that?
Time: Tue Sep 24 10:15:22 2024
User: yes 
Time: Tue Sep 24 10:15:39 2024
Is it ambiguous? False
Time: Tue Sep 24 10:15:39 2024
Is it known? True
Time: Tue Sep 24 10:15:40 2024
var_KnowNo: ['pancakes with maple syrup and berries']
Time: Tue Sep 24 10:15:42 2024
Does it map to the action in var_KnowNo? True
Time: Tue Sep 24 10:15:43 2024
Assistant: Great choice! I will get the task for pancakes with maple syrup and berries started.
Time: Tue Sep 24 10:15:43 2024
Run number: 2 ended at: Tue Sep 24 10:15:43 2024
This run took 5 turns
Total number of LLM calls: 7

Run: 3
Tue Sep 24 10:16:07 2024
Assistant: That's it for this task. Let's go onto the next one... Hello. How can I help you today?
User:  uh can you help me prepare bacon and accent which with peanut butter 
Is it ambiguous? True
Time: Tue Sep 24 10:16:08 2024
Assistant: There are multiple options in my list that might fit your order. It seems like you would like a 'bacon and egg sandwich' but mentioned 'accent', which might be a miscommunication. Additionally, you mentioned peanut butter. 

Could you please clarify if you would like to combine the bacon and egg sandwich with peanut butter, or if you prefer a peanut butter and jelly sandwich instead? Let me know how you would like to proceed!
Time: Tue Sep 24 10:16:09 2024
User: yes I would like to combine get better 
Time: Tue Sep 24 10:16:38 2024
Is it ambiguous? True
Time: Tue Sep 24 10:16:40 2024
Assistant: There are multiple options in my list that might fit your order. To confirm, you would like to combine a 'bacon and egg sandwich' with peanut butter? I can prepare that for you. Please hang tight while I get started!
Time: Tue Sep 24 10:16:41 2024
User: yes generate that sequence for me
Time: Tue Sep 24 10:17:10 2024
Is it ambiguous? False
Time: Tue Sep 24 10:17:11 2024
Is it known? False
Time: Tue Sep 24 10:17:12 2024
Is it within the capabilities of the robot? True
Time: Tue Sep 24 10:17:13 2024
Is a new sequence requested?: True
Time: Tue Sep 24 10:17:13 2024
Is a new sequence actually requested? True
Time: Tue Sep 24 10:17:14 2024
Did it pass the safety check?: True
Explanation: The request involves known ingredients (bacon, egg, and peanut butter) that can be found in the list of known ingredients (No. 3, 10, and 22), making it feasible for me to execute.
Time: Tue Sep 24 10:17:15 2024
The generated sequence is: {"sequence": [{"step": "1", "action": "toast_bread(2)"}, {"step": "2", "action": "fry_eggs(1)"}, {"step": "3", "action": "initiate_sandwich(toasted_bread)"}, {"step": "4", "action": "place_on_sandwich(fried_egg, 1)"}, {"step": "5", "action": "cook_bacon(2)"}, {"step": "6", "action": "place_on_sandwich(fried_bacon, 2)"}, {"step": "7", "action": "spread_on_bread(peanut_butter)"}, {"step": "8", "action": "place_on_sandwich(toasted_bread, 1)"}, {"step": "9", "action": "serve()"}]}
Time: Tue Sep 24 10:17:17 2024
Assistant: Sure, I will perform a new set of steps to fulfill your request for a bacon and egg sandwich with peanut butter. 

First, I will toast two pieces of bread. 
Next, I'll fry one egg. 
Then, I will initiate the sandwich with the toasted bread. 
After placing the fried egg on the sandwich, I'll cook two pieces of bacon. 
Once the bacon is ready, I will add it to the sandwich. 
Following that, I will spread peanut butter on the bread. 
Finally, I'll place the other piece of toasted bread on top and serve the sandwich. 

Does this sequence sound good to you?
Time: Tue Sep 24 10:17:18 2024
var_generated_sequence_name is: bacon and egg sandwich with peanut butter
Time: Tue Sep 24 10:17:18 2024
User:yes 
Time: Tue Sep 24 10:17:54 2024
Assistant: Great then I will now proceed to make the: bacon and egg sandwich with peanut butter
Time: Tue Sep 24 10:17:56 2024
Run number: 3 ended at: Tue Sep 24 10:17:56 2024
This run took 8 turns
Total number of LLM calls: 12

Run: 4
Tue Sep 24 10:18:13 2024
Assistant: That's it for this task. Let's go onto the next one... Hello. How can I help you today?
User: can you help me with washing my laundry 
Is it ambiguous? True
Time: Tue Sep 24 10:18:13 2024
Assistant: I apologize, but washing laundry is outside the scope of my capabilities. However, I can help you with cleaning tasks, such as 'cleaning the living room floor' if that would be helpful. Just let me know!
Time: Tue Sep 24 10:18:14 2024
User:  or can you help me with painting the walls 
Time: Tue Sep 24 10:18:40 2024
Is it ambiguous? True
Time: Tue Sep 24 10:18:42 2024
Assistant: I apologize, but painting walls is also outside the scope of my capabilities. However, I can assist you with cleaning tasks, such as 'cleaning the living room floor'. If that would be helpful, please let me know!
Time: Tue Sep 24 10:18:42 2024
User: skip
Time: Tue Sep 24 10:19:03 2024
Run number: 4 ended at: Tue Sep 24 10:19:03 2024
This run took 6 turns
Total number of LLM calls: 4

Run: 5
Tue Sep 24 10:19:33 2024
Assistant: That's it for this task. Let's go onto the next one... Hello. How can I help you today?
User:   hi can you suggest something healthy uh from the list uh with sausages 
Is it ambiguous? True
Time: Tue Sep 24 10:19:34 2024
Assistant: There are multiple options in my list that might fit your order. Based on your request for something healthy with sausages, I suggest:
1. 'Avocado toast with sausage on the side'
2. 'Full English breakfast' (which includes sausages among other items)

Would you like to go with one of these options?
Time: Tue Sep 24 10:19:35 2024
User: yes I would like to go with the first 1 
Time: Tue Sep 24 10:19:57 2024
Is it ambiguous? False
Time: Tue Sep 24 10:19:59 2024
Is it known? True
Time: Tue Sep 24 10:19:59 2024
var_KnowNo: ['avocado toast with sausage on the side']
Time: Tue Sep 24 10:20:02 2024
Does it map to the action in var_KnowNo? True
Time: Tue Sep 24 10:20:02 2024
Assistant: Great choice! I will get the task for avocado toast with sausage on the side started.
Time: Tue Sep 24 10:20:02 2024
Run number: 5 ended at: Tue Sep 24 10:20:02 2024
This run took 5 turns
Total number of LLM calls: 7

Notes:
  Times: 9:55 - 10:30

  Base: 
    - bad STT at T1
    - interesting T3 request that worked

  BT:
    - T3 had a long break because mic didn't pic up and i had to type in, but request worked and was interesting too.

End of Participant 43 experiment
""" 
parsed_data = parse_text(text)
write_to_excel(parsed_data)

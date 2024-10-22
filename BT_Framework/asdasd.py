import pandas as pd
import re
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

def process_assistant_text(text):
    # Keywords to identify where the JSON sequence starts
    keywords = ["Here�s the new sequence in JSON format:", "JSON format", "Here�s the new sequence I can generate for you"]
    for keyword in keywords:
        if keyword in text:
            # Split the content based on the keyword
            parts = text.split(keyword, 1)
            if len(parts) > 1:
                # Add the specific phrase before the JSON
                modified_text = parts[0].strip() + "\n[this part is not spoken by furhat]\n" + keyword + parts[1].strip()
                return modified_text
    return text

def parse_text(text):
    # Split the text into individual lines
    lines = text.strip().split('\n')

    # Variables to store parsed data
    data = []
    participant_id = None

    # Regular expressions to identify the parts of the text
    id_pattern = re.compile(r'Participant (\d+)')
    user_pattern = re.compile(r'^User:')
    assistant_pattern = re.compile(r'^Assistant:')
    system_keywords = [
        "Is it ambiguous?", "Does it map to the action in var_KnowNo?", 
        "Furhat is being used with Baseline", "var_KnowNo", 
        "Tree is initialized", "Run number", "This run took", "Total number of LLM calls"
    ]

    # Iterate through the lines to categorize and extract information
    current_speaker = None
    current_text = []

    for line in lines:
        # Check for a new participant ID
        match = id_pattern.search(line)
        if match:
            participant_id = match.group(1)
            continue  # Skip to the next line after extracting the participant ID

        if user_pattern.match(line):
            if current_speaker:
                if current_speaker == 'Assistant':
                    # Process and append the entire assistant output before adding it to the data
                    ass_output = "\n".join(current_text).strip()
                    ass_output = process_assistant_text(ass_output)
                    data.append((participant_id, current_speaker, ass_output))
                else:
                    data.append((participant_id, current_speaker, "\n".join(current_text).strip()))
            current_speaker = 'User'
            # Check if the split result has the expected format
            parts = line.split('User: ', 1)
            if len(parts) > 1:
                current_text = [parts[1].strip()]
            else:
                current_text = []
        elif assistant_pattern.match(line):
            if current_speaker:
                if current_speaker == 'Assistant':
                    # Process and append the entire assistant output before adding it to the data
                    ass_output = "\n".join(current_text).strip()
                    ass_output = process_assistant_text(ass_output)
                    data.append((participant_id, current_speaker, ass_output))
                else:
                    data.append((participant_id, current_speaker, "\n".join(current_text).strip()))
            current_speaker = 'Assistant'
            parts = line.split('Assistant: ', 1)
            if len(parts) > 1:
                current_text = [parts[1].strip()]
            else:
                current_text = []
        elif any(keyword in line for keyword in system_keywords) or line.startswith("Time:") or line.startswith("Version:") or line.startswith("Run:"):
            if current_speaker:
                if current_speaker == 'Assistant':
                    ass_output = "\n".join(current_text).strip()
                    ass_output = process_assistant_text(ass_output)
                    data.append((participant_id, current_speaker, ass_output))
                else:
                    data.append((participant_id, current_speaker, "\n".join(current_text).strip()))
            current_speaker = 'System'
            current_text = [line.strip()]
        else:
            current_text.append(line.strip())

    if current_speaker:
        if current_speaker == 'Assistant':
            ass_output = "\n".join(current_text).strip()
            ass_output = process_assistant_text(ass_output)
            data.append((participant_id, current_speaker, ass_output))
        else:
            data.append((participant_id, current_speaker, "\n".join(current_text).strip()))

    return data

def write_to_excel(data, filename='Transkript of Conversation/transcript_parsed.xlsx'):
    # Create a DataFrame
    df = pd.DataFrame(data, columns=['ID', 'Who', 'What'])

    try:
        # Load the existing workbook if it exists
        workbook = load_workbook(filename)
        with pd.ExcelWriter(filename, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            # Add the DataFrame to a new sheet called 'Transcript'
            df.to_excel(writer, index=False, sheet_name='Transcript')

            # Access the newly created sheet for formatting
            worksheet = writer.sheets['Transcript']

            # Apply formatting and merging as needed
            fill_user = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Light green
            fill_assistant = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")  # Light blue
            fill_system = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Light orange

            # Merge cells for the ID column per participant
            start_row = 2
            for row in range(3, len(df) + 2):
                previous_id = worksheet.cell(row=start_row, column=1).value
                current_id = worksheet.cell(row=row, column=1).value

                if current_id != previous_id:
                    worksheet.merge_cells(start_row=start_row, start_column=1, end_row=row - 1, end_column=1)
                    merged_cell = worksheet.cell(row=start_row, column=1)
                    merged_cell.alignment = Alignment(horizontal='center', vertical='top')
                    start_row = row

            if len(df) + 1 > start_row:
                worksheet.merge_cells(start_row=start_row, start_column=1, end_row=len(df) + 1, end_column=1)
                merged_cell = worksheet.cell(row=start_row, column=1)
                merged_cell.alignment = Alignment(horizontal='center', vertical='top')

            # Merge repeated "Who" column values
            start_row = 2
            for row in range(3, len(df) + 2):
                previous_value = worksheet.cell(row=start_row, column=2).value
                current_value = worksheet.cell(row=row, column=2).value

                if current_value != previous_value:
                    if row - 1 > start_row:
                        worksheet.merge_cells(start_row=start_row, start_column=2, end_row=row - 1, end_column=2)
                        merged_cell = worksheet.cell(row=start_row, column=2)
                        merged_cell.value = f"{previous_value}:"
                        merged_cell.alignment = Alignment(vertical='top')
                    start_row = row

            if len(df) + 1 > start_row:
                worksheet.merge_cells(start_row=start_row, start_column=2, end_row=len(df) + 1, end_column=2)
                merged_cell = worksheet.cell(row=start_row, column=2)
                merged_cell.alignment = Alignment(vertical='top')

            for row in range(2, len(df) + 2):
                who_value = worksheet.cell(row=row, column=2).value
                who_cell = worksheet.cell(row=row, column=2)
                what_cell = worksheet.cell(row=row, column=3)

                cell_coord = f"B{row}"
                merged = any(cell_coord in merged_range for merged_range in worksheet.merged_cells.ranges)

                if not merged:
                    who_cell.value = f"{who_value}:"

                if who_value == 'User' or who_value == 'User:':
                    who_cell.fill = fill_user
                elif who_value == 'Assistant' or who_value == 'Assistant:':
                    who_cell.fill = fill_assistant
                elif who_value == 'System' or who_value == 'System:':
                    who_cell.fill = fill_system

                what_cell.alignment = Alignment(wrap_text=True, vertical='top')
                who_cell.alignment = Alignment(vertical='top')

            for col in range(1, 4):
                max_length = 0
                col_letter = get_column_letter(col)
                for cell in worksheet[col_letter]:
                    try:
                        max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                worksheet.column_dimensions[col_letter].width = max_length + 5

    except FileNotFoundError:
        # If the file does not exist, create a new one
        with pd.ExcelWriter(filename, engine='openpyxl', mode='w') as writer:
            df.to_excel(writer, index=False, sheet_name='Transcript')

# Example usage
text = """
Participant 01
Run ID A
Version: 1.3
Sun Sep  8 11:14:43 2024
Furhat is being used with Baseline
User: I would like a bacon and egg sandwich please 
Time: Sun Sep  8 11:14:51 2024
Assistant: Ok, I will now proceed with the sequence: '1', 'bacon and egg sandwich'.
Time: Sun Sep  8 11:14:52 2024
Run ended at: 1
Sun Sep  8 11:14:52 2024
This run took 1 turns
Total number of LLM calls: 1

Version: 1.3
Sun Sep  8 11:14:57 2024
Furhat is being used with Baseline
User: you get me a bacon and egg sandwich 
Time: Sun Sep  8 11:15:13 2024
Assistant: Ok I will now proceed with the sequence: '1', 'bacon and egg sandwich'.
Time: Sun Sep  8 11:15:13 2024
Run ended at: 2
Sun Sep  8 11:15:13 2024
This run took 2 turns
Total number of LLM calls: 1

Version: 1.3
Sun Sep  8 11:15:21 2024
Furhat is being used with Baseline
User: get a bacon and egg sandwich please 
Time: Sun Sep  8 11:16:13 2024
Assistant: Ok I will now proceed with the sequence: '1', 'bacon and egg sandwich'.
Time: Sun Sep  8 11:16:14 2024
Run ended at: 3
Sun Sep  8 11:16:14 2024
This run took 3 turns
Total number of LLM calls: 1

Version: 1.3
Sun Sep  8 11:16:32 2024
Furhat is being used with Baseline
User: can I get a bacon and egg sandwich please 
Time: Sun Sep  8 11:16:41 2024
Assistant: Ok, I will now proceed with the sequence: '1', 'bacon and egg sandwich'.
Time: Sun Sep  8 11:16:41 2024
Run ended at: 1
Sun Sep  8 11:16:41 2024
This run took 1 turns
Total number of LLM calls: 1

Version: 1.3
Sun Sep  8 11:16:47 2024
Furhat is being used with Baseline
User: you suggest me something sweet 
Time: Sun Sep  8 11:16:58 2024
Assistant: I understand you're looking for a sweet suggestion. Here are a few options from my capabilities:

1. **Pancakes with Maple Syrup and Berries** - Fluffy pancakes served with sweet maple syrup and fresh berries.
2. **Peanut Butter and Jelly Sandwich** - A classic sandwich with creamy peanut butter and sweet jelly.

Please let me know if you would like me to prepare one of these options!
Time: Sun Sep  8 11:16:59 2024
User: can I get the peanut butter and jelly sandwich 
Time: Sun Sep  8 11:17:28 2024
Assistant: Ok, I will now proceed with the sequence: '3', 'peanut butter and jelly sandwich'. 
Time: Sun Sep  8 11:17:29 2024
Run ended at: 2
Sun Sep  8 11:17:29 2024
This run took 2 turns
Total number of LLM calls: 2

Version: 1.3
Sun Sep  8 11:17:35 2024
Furhat is being used with Baseline
User: I get uh full English breakfast without the sausages 
Time: Sun Sep  8 11:17:49 2024
Assistant: Alright I will create that sequence for you.
Assistant: The sequence is: I cannot provide the full English breakfast without sausages as it deviates from the predefined sequence. However, I can proceed with a modified sequence that maintains all the other elements of the full English breakfast, excluding the sausages. 

Here�s the new sequence I can generate for you:

```json
{
  "task": "full English breakfast without sausages",
  "ID": "6_modified",
  "sequence": [
    {"step": "1", "action": "cook_bacon(2)"},
    {"step": "2", "action": "toast_bread(2)"},
    {"step": "3", "action": "grill_tomatoes(1)"},
    {"step": "4", "action": "grill_mushrooms(3)"},
    {"step": "5", "action": "heat_beans(1)"},
    {"step": "6", "action": "fry_eggs(1)"},
    {"step": "7", "action": "place_on_plate(fried_bacon, 2)"},
    {"step": "8", "action": "place_on_plate(toasted_bread, 2)"},
    {"step": "9", "action": "place_on_plate(grilled_tomatoes, 1)"},
    {"step": "10", "action": "place_on_plate(grilled_mushrooms, 3)"},
    {"step": "11", "action": "place_on_plate(heated_beans, 1)"},
    {"step": "12", "action": "place_on_plate(fried_egg, 1)"},
    {"step": "13", "action": "serve()"}
  ]
}
```

If you'd like me to proceed with this modified sequence, just let me know!
Time: Sun Sep  8 11:17:52 2024
Run ended at: 3
Sun Sep  8 11:17:52 2024
This run took 3 turns
Total number of LLM calls: 1

Version: 1.3
Sun Sep  8 11:18:18 2024
Furhat is being used with Baseline
User: drive me to my next location 
Time: Sun Sep  8 11:18:29 2024
Assistant: I'm sorry, but driving or transportation is beyond my current capabilities. I specialize in food preparation and cleaning tasks. However, if you need assistance with a meal or cleaning, I can certainly help with that! Please let me know what you'd like.
Time: Sun Sep  8 11:18:30 2024
User: you could help me clean 
Time: Sun Sep  8 11:18:51 2024
Assistant: Great! I can assist you with cleaning tasks. I can perform the following task:

1. Clean the living room floor.

Please confirm if you would like me to proceed with that task, or let me know if you have another specific cleaning request!
Time: Sun Sep  8 11:18:52 2024
User: yes proceed with the living room cleaning 
Time: Sun Sep  8 11:19:12 2024
Assistant: Ok, I will now proceed with the sequence: '10', 'clean living room floor'. 

I'll start by retrieving the cleaning supplies.
Time: Sun Sep  8 11:19:13 2024
Run ended at: 4
Sun Sep  8 11:19:13 2024
This run took 4 turns
Total number of LLM calls: 3

Version: 1.3
Sun Sep  8 11:21:43 2024
Tree is initialized and furhat is used

Assistant: Hello! I am Gregory. How can I help you today?
User: can I get a bacon and egg sandwich please 
Is it ambiguous? False
Time: Sun Sep  8 11:21:44 2024
Is it known? True
Time: Sun Sep  8 11:21:45 2024
var_KnowNo: ['bacon and egg sandwich']
Time: Sun Sep  8 11:21:46 2024
Does it map to the action in var_KnowNo? True
Time: Sun Sep  8 11:21:47 2024
Assistant: Great choice! I will execute the sequence for bacon and egg sandwich.
Time: Sun Sep  8 11:21:47 2024
Run ended at: 1
Sun Sep  8 11:21:47 2024
This run took 3 turns
Total number of LLM calls: 5

Run: 2
Sun Sep  8 11:22:02 2024
Assistant: Let's go again... Hello. How can I help you today?
User: can you suggest me something savory 
Is it ambiguous? True
Time: Sun Sep  8 11:22:03 2024
Assistant: There are multiple options in the menu that might fit your order. Here are a few savory options you might enjoy:
1. A 'full English breakfast', which includes a variety of cooked items.
2. 'Avocado toast with sausage on the side' for a delicious combination.
3. A 'bean and cheese quesadilla' for a tasty snack.

Let me know if any of these options sound good to you!
Time: Sun Sep  8 11:22:04 2024
User: could I get the bean and cheese quesadilla 
Time: Sun Sep  8 11:22:33 2024
Is it ambiguous? False
Time: Sun Sep  8 11:22:34 2024
Is it known? True
Time: Sun Sep  8 11:22:34 2024
var_KnowNo: ['bean and cheese quesadilla']
Time: Sun Sep  8 11:22:36 2024
Does it map to the action in var_KnowNo? True
Time: Sun Sep  8 11:22:36 2024
Assistant: Great choice! I will execute the sequence for bean and cheese quesadilla.
Time: Sun Sep  8 11:22:36 2024
Run ended at: 2
Sun Sep  8 11:22:36 2024
This run took 5 turns
Total number of LLM calls: 7

Run: 3
Sun Sep  8 11:22:55 2024
Assistant: Let's go again... Hello. How can I help you today?
User: can I get the grilled tomatoes and mushroom bruschetta without the tomatoes 
Is it ambiguous? False
Time: Sun Sep  8 11:22:56 2024
Is it known? False
Time: Sun Sep  8 11:22:56 2024
Is it within the capabilities of the robot? True
Time: Sun Sep  8 11:22:57 2024
Is a new sequence requested?: True
Time: Sun Sep  8 11:22:57 2024
Is a new sequence actually requested? True
Time: Sun Sep  8 11:22:58 2024
Did it pass the safety check?: True
Explanation: The request involves known ingredients (mushrooms and bread) that can be found in the list of known ingredients (No. 1 and 6), and modifying the dish by removing tomatoes is reasonable, making it feasible for me to execute.
Time: Sun Sep  8 11:22:59 2024
The generated sequence is: {"sequence": [{"step": "1", "action": "grill_mushrooms(2)"}, {"step": "2", "action": "toast_bread(2)"}, {"step": "3", "action": "initiate_sandwich(toasted_bread)"}, {"step": "4", "action": "place_on_sandwich(grilled_mushrooms, 1)"}, {"step": "5", "action": "initiate_sandwich(toasted_bread)"}, {"step": "6", "action": "place_on_sandwich(grilled_mushrooms, 1)"}, {"step": "7", "action": "serve()"}]}
Time: Sun Sep  8 11:23:00 2024
Assistant: Sure, I will generate a new sequence to fulfill your request. First, I'll grill two mushrooms. 
Then, I'll toast two pieces of bread. 
Next, I'll initiate a sandwich using the toasted bread. 
I'll place one grilled mushroom on the sandwich. 
I'll repeat the sandwich initiation for the second piece of toasted bread. 
Afterward, I'll place one more grilled mushroom on that sandwich. 
Finally, I'll serve the finished bruschetta. 
Does this sequence sound good to you?
Time: Sun Sep  8 11:23:02 2024
var_generated_sequence_name is: grilled mushroom bruschetta without tomatoes
Time: Sun Sep  8 11:23:02 2024
User:yes 
Time: Sun Sep  8 11:23:34 2024
Assistant: Great then I will now proceed to make the: grilled mushroom bruschetta without tomatoes
Time: Sun Sep  8 11:23:35 2024
Run ended at: 3
Sun Sep  8 11:23:35 2024
This run took 4 turns
Total number of LLM calls: 8

Run: 4
Sun Sep  8 11:23:51 2024
Assistant: Let's go again... Hello. How can I help you today?
User: we play basketball with me 
Is it ambiguous? True
Time: Sun Sep  8 11:23:51 2024
Assistant: I apologize, but playing basketball is outside the scope of my capabilities. However, I can assist you with cooking or cleaning. If you're hungry after playing, I can suggest some meals like a 'bacon and egg sandwich' or 'vegetable stir fry with rice'. Please let me know how I can assist you!
Time: Sun Sep  8 11:23:53 2024
User: you could make a pancakes with make maple syrup 
Time: Sun Sep  8 11:24:18 2024
Is it ambiguous? True
Time: Sun Sep  8 11:24:18 2024
Assistant: There are multiple options in the menu that might fit your order. I understand you'd like 'pancakes with maple syrup'. Would you like me to prepare that for you? If you have any specific quantity in mind, please let me know!
Time: Sun Sep  8 11:24:20 2024
User: no that's the normal 1 
Time: Sun Sep  8 11:24:38 2024
Is it ambiguous? True
Time: Sun Sep  8 11:24:38 2024
Assistant: Great! I will proceed with preparing one serving of 'pancakes with maple syrup and berries'. Just a moment while I get that ready for you!
Time: Sun Sep  8 11:24:39 2024
User: esc
Time: Sun Sep  8 11:24:52 2024
Run ended at: 4
Sun Sep  8 11:24:52 2024
This run took 8 turns
Total number of LLM calls: 6



Notes:
- Time: 11:10 - 11:30
- Had to restart becuase "lets go again ... Hello how can i help you today" made participant restart from task 1, had to step in to clarify and restart. Later code was adjusted to make clear task is over and moving on to next task with: "That's it for this task. Let's go onto the next one... Hello. How can I help you today?".

End of Participant 01 experiment
""" 
parsed_data = parse_text(text)
write_to_excel(parsed_data)
